"""
Excel Bid Generator Module

This module provides functionality to create professional Excel bid documents
with multiple sheets, professional formatting, and comprehensive styling.
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelBidGenerator:
    """
    Professional Excel Bid Document Generator
    
    Creates comprehensive Excel bid documents with multiple sheets:
    - Executive Summary
    - Line Items Detail  
    - CalTrans Analysis
    
    Features professional formatting, company branding, and comprehensive styling.
    """
    
    def __init__(self, company_name: str = "Zenlytic Solutions", 
                 company_logo_path: Optional[str] = None):
        """
        Initialize the Excel Bid Generator
        
        Args:
            company_name: Company name for branding
            company_logo_path: Optional path to company logo
        """
        self.company_name = company_name
        self.company_logo_path = company_logo_path
        
        # Define color scheme
        self.colors = {
            'header_fill': '1F4E79',  # Dark blue
            'header_font': 'FFFFFF',  # White
            'subheader_fill': '4472C4',  # Medium blue
            'subheader_font': 'FFFFFF',  # White
            'accent_fill': 'D9E2F3',  # Light blue
            'border': '000000',  # Black
            'success_fill': 'C6EFCE',  # Light green
            'warning_fill': 'FFEB9C',  # Light yellow
            'error_fill': 'FFC7CE'  # Light red
        }
        
        # Define styles
        self.styles = {}
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup named styles for consistent formatting"""
        try:
            # Header style
            header_style = NamedStyle(name="header_style")
            header_style.font = Font(
                name='Arial', 
                size=12, 
                bold=True, 
                color=self.colors['header_font']
            )
            header_style.fill = PatternFill(
                start_color=self.colors['header_fill'],
                end_color=self.colors['header_fill'],
                fill_type='solid'
            )
            header_style.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            header_style.border = Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            )
            self.styles['header'] = header_style
            
            # Subheader style
            subheader_style = NamedStyle(name="subheader_style")
            subheader_style.font = Font(
                name='Arial', 
                size=11, 
                bold=True, 
                color=self.colors['subheader_font']
            )
            subheader_style.fill = PatternFill(
                start_color=self.colors['subheader_fill'],
                end_color=self.colors['subheader_fill'],
                fill_type='solid'
            )
            subheader_style.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            subheader_style.border = Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            )
            self.styles['subheader'] = subheader_style
            
            # Data style
            data_style = NamedStyle(name="data_style")
            data_style.font = Font(name='Arial', size=10)
            data_style.alignment = Alignment(
                horizontal='left',
                vertical='center',
                wrap_text=True
            )
            data_style.border = Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            )
            self.styles['data'] = data_style
            
            # Currency style
            currency_style = NamedStyle(name="currency_style")
            currency_style.font = Font(name='Arial', size=10)
            currency_style.alignment = Alignment(
                horizontal='right',
                vertical='center'
            )
            currency_style.border = Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            )
            currency_style.number_format = '$#,##0.00'
            self.styles['currency'] = currency_style
            
            # Number style
            number_style = NamedStyle(name="number_style")
            number_style.font = Font(name='Arial', size=10)
            number_style.alignment = Alignment(
                horizontal='right',
                vertical='center'
            )
            number_style.border = Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            )
            number_style.number_format = '#,##0'
            self.styles['number'] = number_style
            
        except Exception as e:
            logger.error(f"Error setting up styles: {e}")
            raise
    
    def create_professional_bid(self, bid_data: Dict[str, Any]) -> bytes:
        """
        Create a professional Excel bid document
        
        Args:
            bid_data: Dictionary containing bid information including:
                - project_info: Project details
                - line_items: List of bid line items
                - caltrans_analysis: CalTrans analysis results
                - pricing_summary: Pricing calculations
                
        Returns:
            bytes: Excel file as bytes for download
        """
        try:
            logger.info("Creating professional Excel bid document")
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self.create_summary_sheet(wb, bid_data)
            self.create_line_items_sheet(wb, bid_data)
            self.create_analysis_sheet(wb, bid_data)
            
            # Save to bytes
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            logger.info("Excel bid document created successfully")
            return excel_file.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating Excel bid document: {e}")
            raise
    
    def generate_comprehensive_bid_excel(
        self, 
        bid_data: Dict[str, Any], 
        analysis_results: Dict[str, Any], 
        output_path: str
    ) -> None:
        """
        Generate comprehensive Excel bid document with multi-document analysis.
        
        Args:
            bid_data: Dictionary containing bid information
            analysis_results: Combined analysis results from multiple documents
            output_path: Path to save the Excel file
        """
        try:
            logger.info("Creating comprehensive Excel bid document")
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create comprehensive sheets
            self.create_comprehensive_summary_sheet(wb, bid_data, analysis_results)
            self.create_comprehensive_line_items_sheet(wb, bid_data, analysis_results)
            self.create_cross_reference_sheet(wb, analysis_results)
            self.create_document_analysis_sheet(wb, analysis_results)
            self.create_analysis_sheet(wb, bid_data)
            
            # Save to file
            wb.save(output_path)
            logger.info(f"Comprehensive Excel bid saved to: {output_path}")
            
        except Exception as e:
            logger.error(f"Error creating comprehensive Excel bid: {e}")
            raise
    
    def create_comprehensive_summary_sheet(
        self, 
        wb: Workbook, 
        bid_data: Dict[str, Any], 
        analysis_results: Dict[str, Any]
    ) -> None:
        """Create comprehensive summary sheet with multi-document analysis"""
        
        ws = wb.create_sheet("Executive Summary")
        
        # Company header
        self._add_company_header(ws, 1)
        
        # Project information
        row = 5
        row = self._add_section_header(ws, row, "Project Information")
        
        ws.cell(row=row, column=1, value="Project Name:")
        ws.cell(row=row, column=2, value=bid_data.get('project_name', 'N/A'))
        row += 1
        
        ws.cell(row=row, column=1, value="Project Number:")
        ws.cell(row=row, column=2, value=bid_data.get('project_number', 'N/A'))
        row += 1
        
        ws.cell(row=row, column=1, value="Bid Date:")
        ws.cell(row=row, column=2, value=bid_data.get('bid_date', 'N/A'))
        row += 2
        
        # Analysis summary
        row = self._add_section_header(ws, row, "Comprehensive Analysis Summary")
        
        ws.cell(row=row, column=1, value="Documents Analyzed:")
        ws.cell(row=row, column=2, value=analysis_results.get('total_documents', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Terms Found:")
        ws.cell(row=row, column=2, value=analysis_results.get('total_terms', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Quantities Extracted:")
        ws.cell(row=row, column=2, value=analysis_results.get('total_quantities', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Alerts Generated:")
        ws.cell(row=row, column=2, value=analysis_results.get('total_alerts', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Confidence Score:")
        ws.cell(row=row, column=2, value=f"{analysis_results.get('confidence_score', 0):.1%}")
        row += 2
        
        # Financial summary
        row = self._add_section_header(ws, row, "Financial Summary")
        
        pricing = bid_data.get('pricing_summary', {})
        
        ws.cell(row=row, column=1, value="Subtotal:")
        ws.cell(row=row, column=2, value=f"${pricing.get('subtotal', 0):,.2f}")
        row += 1
        
        ws.cell(row=row, column=1, value="Markup:")
        ws.cell(row=row, column=2, value=f"${pricing.get('markup_amount', 0):,.2f}")
        row += 1
        
        ws.cell(row=row, column=1, value="Delivery Fee:")
        ws.cell(row=row, column=2, value=f"${pricing.get('delivery_fee', 0):,.2f}")
        row += 1
        
        ws.cell(row=row, column=1, value="Waste Adjustments:")
        ws.cell(row=row, column=2, value=f"${pricing.get('waste_adjustments', 0):,.2f}")
        row += 1
        
        ws.cell(row=row, column=1, value="Total:")
        ws.cell(row=row, column=2, value=f"${pricing.get('total', 0):,.2f}")
        row += 2
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def create_comprehensive_line_items_sheet(
        self, 
        wb: Workbook, 
        bid_data: Dict[str, Any], 
        analysis_results: Dict[str, Any]
    ) -> None:
        """Create comprehensive line items sheet with cross-reference information"""
        
        ws = wb.create_sheet("Line Items - Comprehensive")
        
        # Company header
        self._add_company_header(ws, 1)
        
        # Section header
        row = self._add_section_header(ws, 5, "Comprehensive Line Items")
        
        # Headers
        headers = [
            "Item #", "Description", "CalTrans Term", "Quantity", "Unit", 
            "Unit Price", "Total Price", "Waste Factor", "Confidence", "Cross-Reference Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = self.styles['header']
        
        row += 1
        
        # Line items data
        line_items = bid_data.get('line_items', [])
        
        for item in line_items:
            ws.cell(row=row, column=1, value=item.get('item_number', ''))
            ws.cell(row=row, column=2, value=item.get('description', ''))
            ws.cell(row=row, column=3, value=item.get('caltrans_term', ''))
            ws.cell(row=row, column=4, value=item.get('quantity', 0))
            ws.cell(row=row, column=5, value=item.get('unit', ''))
            ws.cell(row=row, column=6, value=f"${item.get('unit_price', 0):,.2f}")
            ws.cell(row=row, column=7, value=f"${item.get('total_price', 0):,.2f}")
            ws.cell(row=row, column=8, value=f"{item.get('waste_factor', 0):.1%}")
            ws.cell(row=row, column=9, value=f"{item.get('confidence', 0):.1%}")
            ws.cell(row=row, column=10, value=item.get('notes', ''))
            row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def create_cross_reference_sheet(self, wb: Workbook, analysis_results: Dict[str, Any]) -> None:
        """Create cross-reference analysis sheet"""
        
        ws = wb.create_sheet("Cross-Reference Analysis")
        
        # Company header
        self._add_company_header(ws, 1)
        
        # Section header
        row = self._add_section_header(ws, 5, "Cross-Reference Analysis")
        
        # Term consistency
        row = self._add_section_header(ws, row, "Term Consistency Across Documents")
        
        headers = ["Term", "Found In", "Consistency", "Document Count"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = self.styles['header']
        
        row += 1
        
        cross_refs = analysis_results.get('cross_references', {})
        term_matches = cross_refs.get('term_matches', [])
        
        for match in term_matches:
            ws.cell(row=row, column=1, value=match.get('term', ''))
            ws.cell(row=row, column=2, value=', '.join(match.get('found_in', [])))
            ws.cell(row=row, column=3, value=match.get('consistency', '').title())
            ws.cell(row=row, column=4, value=len(match.get('found_in', [])))
            row += 1
        
        row += 2
        
        # Quantity discrepancies
        row = self._add_section_header(ws, row, "Quantity Discrepancies")
        
        headers = ["Specification Value", "Bid Form Value", "Unit", "Difference %"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = self.styles['header']
        
        row += 1
        
        discrepancies = cross_refs.get('quantity_discrepancies', [])
        
        for disc in discrepancies:
            ws.cell(row=row, column=1, value=disc.get('specification_value', 0))
            ws.cell(row=row, column=2, value=disc.get('bid_form_value', 0))
            ws.cell(row=row, column=3, value=disc.get('unit', ''))
            ws.cell(row=row, column=4, value=f"{disc.get('difference_percent', 0):.1f}%")
            row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def create_document_analysis_sheet(self, wb: Workbook, analysis_results: Dict[str, Any]) -> None:
        """Create document analysis summary sheet"""
        
        ws = wb.create_sheet("Document Analysis")
        
        # Company header
        self._add_company_header(ws, 1)
        
        # Section header
        row = self._add_section_header(ws, 5, "Document Analysis Summary")
        
        # Document coverage
        row = self._add_section_header(ws, row, "Document Coverage Analysis")
        
        headers = ["Document Type", "Pages Analyzed", "Terms Found", "Quantities Found", "Confidence Score", "Quality Score"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = self.styles['header']
        
        row += 1
        
        cross_refs = analysis_results.get('cross_references', {})
        doc_coverage = cross_refs.get('document_coverage', {})
        
        for doc_type, coverage in doc_coverage.items():
            ws.cell(row=row, column=1, value=doc_type.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=coverage.get('pages_analyzed', 0))
            ws.cell(row=row, column=3, value=coverage.get('terms_found', 0))
            ws.cell(row=row, column=4, value=coverage.get('quantities_found', 0))
            ws.cell(row=row, column=5, value=f"{coverage.get('confidence_score', 0):.1%}")
            ws.cell(row=row, column=6, value=f"{coverage.get('quality_score', 0):.1%}")
            row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def create_summary_sheet(self, wb: Workbook, bid_data: Dict[str, Any]) -> None:
        """
        Create Executive Summary sheet
        
        Args:
            wb: Workbook object
            bid_data: Bid data dictionary
        """
        try:
            ws = wb.create_sheet("Executive Summary")
            
            # Get data
            project_info = bid_data.get('project_info', {})
            pricing_summary = bid_data.get('pricing_summary', {})
            
            # Company header
            self._add_company_header(ws, 1)
            
            # Project information section
            current_row = 4
            current_row = self._add_section_header(ws, current_row, "Project Information")
            
            project_fields = [
                ("Project Name", project_info.get('project_name', 'N/A')),
                ("Project Number", project_info.get('project_number', 'N/A')),
                ("Bid Date", datetime.now().strftime('%Y-%m-%d')),
                ("Company", self.company_name),
                ("Contact Person", project_info.get('contact_person', 'N/A')),
                ("Phone", project_info.get('phone', 'N/A')),
                ("Email", project_info.get('email', 'N/A'))
            ]
            
            for field_name, field_value in project_fields:
                ws.cell(row=current_row, column=1, value=field_name).style = self.styles['data']
                ws.cell(row=current_row, column=2, value=field_value).style = self.styles['data']
                current_row += 1
            
            # Pricing summary section
            current_row += 2
            current_row = self._add_section_header(ws, current_row, "Pricing Summary")
            
            pricing_fields = [
                ("Subtotal", pricing_summary.get('subtotal', 0)),
                ("Tax Rate", f"{pricing_summary.get('tax_rate', 0):.2%}"),
                ("Tax Amount", pricing_summary.get('tax_amount', 0)),
                ("Shipping", pricing_summary.get('shipping', 0)),
                ("Handling", pricing_summary.get('handling', 0)),
                ("Total", pricing_summary.get('total', 0))
            ]
            
            for field_name, field_value in pricing_fields:
                ws.cell(row=current_row, column=1, value=field_name).style = self.styles['data']
                if isinstance(field_value, (int, float)) and field_name != "Tax Rate":
                    ws.cell(row=current_row, column=2, value=field_value).style = self.styles['currency']
                else:
                    ws.cell(row=current_row, column=2, value=field_value).style = self.styles['data']
                current_row += 1
            
            # Add total row with emphasis
            total_row = current_row - 1
            ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=total_row, column=2).font = Font(bold=True, size=12)
            
            # Bid notes section
            current_row += 2
            current_row = self._add_section_header(ws, current_row, "Bid Notes")
            
            notes = bid_data.get('notes', 'No additional notes provided.')
            ws.cell(row=current_row, column=1, value="Notes").style = self.styles['data']
            ws.cell(row=current_row, column=2, value=notes).style = self.styles['data']
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {e}")
            raise
    
    def create_line_items_sheet(self, wb: Workbook, bid_data: Dict[str, Any]) -> None:
        """
        Create Line Items Detail sheet
        
        Args:
            wb: Workbook object
            bid_data: Bid data dictionary
        """
        try:
            ws = wb.create_sheet("Line Items Detail")
            
            # Get line items
            line_items = bid_data.get('line_items', [])
            
            # Company header
            self._add_company_header(ws, 1)
            
            # Table header
            current_row = 4
            headers = [
                "Item #", "SKU", "Description", "Quantity", 
                "Unit Price", "Extended Price", "Notes"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.style = self.styles['header']
            
            # Add data rows
            current_row += 1
            for i, item in enumerate(line_items, 1):
                row_data = [
                    i,  # Item #
                    item.get('sku', 'N/A'),
                    item.get('description', 'N/A'),
                    item.get('quantity', 0),
                    item.get('unit_price', 0),
                    item.get('extended_price', 0),
                    item.get('notes', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    
                    # Apply appropriate styling
                    if col in [4]:  # Quantity
                        cell.style = self.styles['number']
                    elif col in [5, 6]:  # Unit Price, Extended Price
                        cell.style = self.styles['currency']
                    else:
                        cell.style = self.styles['data']
                
                current_row += 1
            
            # Add totals row
            if line_items:
                total_row = current_row
                ws.cell(row=total_row, column=1, value="TOTAL").style = self.styles['header']
                ws.cell(row=total_row, column=2, value="").style = self.styles['header']
                ws.cell(row=total_row, column=3, value="").style = self.styles['header']
                ws.cell(row=total_row, column=4, value="").style = self.styles['header']
                ws.cell(row=total_row, column=5, value="").style = self.styles['header']
                
                # Calculate total extended price
                total_extended = sum(item.get('extended_price', 0) for item in line_items)
                ws.cell(row=total_row, column=6, value=total_extended).style = self.styles['currency']
                ws.cell(row=total_row, column=7, value="").style = self.styles['header']
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating line items sheet: {e}")
            raise
    
    def create_analysis_sheet(self, wb: Workbook, bid_data: Dict[str, Any]) -> None:
        """
        Create CalTrans Analysis sheet
        
        Args:
            wb: Workbook object
            bid_data: Bid data dictionary
        """
        try:
            ws = wb.create_sheet("CalTrans Analysis")
            
            # Get analysis data
            caltrans_analysis = bid_data.get('caltrans_analysis', {})
            
            # Company header
            self._add_company_header(ws, 1)
            
            # Analysis summary section
            current_row = 4
            current_row = self._add_section_header(ws, current_row, "Analysis Summary")
            
            summary_fields = [
                ("Total Terms Found", caltrans_analysis.get('total_terms', 0)),
                ("Matched Products", caltrans_analysis.get('matched_products', 0)),
                ("Unmatched Terms", caltrans_analysis.get('unmatched_terms', 0)),
                ("Confidence Score", f"{caltrans_analysis.get('confidence_score', 0):.1%}"),
                ("Analysis Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            ]
            
            for field_name, field_value in summary_fields:
                ws.cell(row=current_row, column=1, value=field_name).style = self.styles['data']
                if isinstance(field_value, (int, float)) and field_name != "Confidence Score":
                    ws.cell(row=current_row, column=2, value=field_value).style = self.styles['number']
                else:
                    ws.cell(row=current_row, column=2, value=field_value).style = self.styles['data']
                current_row += 1
            
            # Terms found section
            current_row += 2
            current_row = self._add_section_header(ws, current_row, "Terms Found")
            
            terms_found = caltrans_analysis.get('terms_found', [])
            if terms_found:
                headers = ["Term", "Quantity", "Unit", "Confidence", "Matched SKU", "Notes"]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.style = self.styles['subheader']
                
                current_row += 1
                
                for term in terms_found:
                    row_data = [
                        term.get('term', 'N/A'),
                        term.get('quantity', 0),
                        term.get('unit', 'N/A'),
                        f"{term.get('confidence', 0):.1%}",
                        term.get('matched_sku', 'N/A'),
                        term.get('notes', '')
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        
                        if col == 2:  # Quantity
                            cell.style = self.styles['number']
                        elif col == 4:  # Confidence
                            cell.style = self.styles['data']
                        else:
                            cell.style = self.styles['data']
                    
                    current_row += 1
            
            # Alerts section
            current_row += 2
            current_row = self._add_section_header(ws, current_row, "Alerts & Warnings")
            
            alerts = caltrans_analysis.get('alerts', [])
            if alerts:
                headers = ["Type", "Message", "Severity", "Recommendation"]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.style = self.styles['subheader']
                
                current_row += 1
                
                for alert in alerts:
                    row_data = [
                        alert.get('type', 'N/A'),
                        alert.get('message', 'N/A'),
                        alert.get('severity', 'N/A'),
                        alert.get('recommendation', 'N/A')
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.style = self.styles['data']
                    
                    current_row += 1
            else:
                ws.cell(row=current_row, column=1, value="No alerts or warnings found.").style = self.styles['data']
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating analysis sheet: {e}")
            raise
    
    def _add_company_header(self, ws: Worksheet, row: int) -> None:
        """
        Add company header to worksheet
        
        Args:
            ws: Worksheet object
            row: Row number to add header
        """
        try:
            # Company name
            ws.cell(row=row, column=1, value=self.company_name)
            ws.cell(row=row, column=1).font = Font(name='Arial', size=16, bold=True, color=self.colors['header_fill'])
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Merge cells for company name
            ws.merge_cells(f'A{row}:G{row}')
            
            # Subtitle
            ws.cell(row=row+1, column=1, value="Professional Bid Document")
            ws.cell(row=row+1, column=1).font = Font(name='Arial', size=12, italic=True, color=self.colors['subheader_fill'])
            ws.cell(row=row+1, column=1).alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{row+1}:G{row+1}')
            
        except Exception as e:
            logger.error(f"Error adding company header: {e}")
            raise
    
    def _add_section_header(self, ws: Worksheet, row: int, title: str) -> int:
        """
        Add section header to worksheet
        
        Args:
            ws: Worksheet object
            row: Row number to add header
            title: Section title
            
        Returns:
            int: Next row number
        """
        try:
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=1).font = Font(name='Arial', size=14, bold=True, color=self.colors['header_font'])
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=self.colors['header_fill'],
                end_color=self.colors['header_fill'],
                fill_type='solid'
            )
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
            
            # Merge cells for section header
            ws.merge_cells(f'A{row}:G{row}')
            
            return row + 1
            
        except Exception as e:
            logger.error(f"Error adding section header: {e}")
            raise
    
    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """
        Auto-adjust column widths based on content
        
        Args:
            ws: Worksheet object
        """
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set minimum and maximum widths
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error auto-adjusting columns: {e}")
            # Don't raise exception for column adjustment errors
    
    def add_logo(self, ws: Worksheet, logo_path: str, row: int = 1, col: int = 1) -> None:
        """
        Add company logo to worksheet (if logo path is provided)
        
        Args:
            ws: Worksheet object
            logo_path: Path to logo image
            row: Row position
            col: Column position
        """
        try:
            if logo_path and self.company_logo_path:
                from openpyxl.drawing.image import Image
                img = Image(self.company_logo_path)
                img.width = 100
                img.height = 50
                ws.add_image(img, f'{get_column_letter(col)}{row}')
                
        except Exception as e:
            logger.warning(f"Could not add logo: {e}")
            # Don't raise exception for logo errors


def create_sample_bid_data() -> Dict[str, Any]:
    """
    Create sample bid data for testing
    
    Returns:
        Dict containing sample bid data
    """
    return {
        'project_info': {
            'project_name': 'CalTrans Highway Maintenance Project',
            'project_number': 'CT-2024-001',
            'contact_person': 'John Smith',
            'phone': '(555) 123-4567',
            'email': 'john.smith@company.com'
        },
        'line_items': [
            {
                'sku': 'CT-001',
                'description': 'Traffic Cone - 28" Orange',
                'quantity': 100,
                'unit_price': 12.50,
                'extended_price': 1250.00,
                'notes': 'High visibility safety equipment'
            },
            {
                'sku': 'CT-002',
                'description': 'Barricade - Type III',
                'quantity': 25,
                'unit_price': 45.00,
                'extended_price': 1125.00,
                'notes': 'Heavy-duty construction barrier'
            },
            {
                'sku': 'CT-003',
                'description': 'Safety Vest - Class 2',
                'quantity': 50,
                'unit_price': 18.75,
                'extended_price': 937.50,
                'notes': 'ANSI compliant safety vest'
            }
        ],
        'caltrans_analysis': {
            'total_terms': 15,
            'matched_products': 12,
            'unmatched_terms': 3,
            'confidence_score': 0.85,
            'terms_found': [
                {
                    'term': 'Traffic Cone',
                    'quantity': 100,
                    'unit': 'EA',
                    'confidence': 0.95,
                    'matched_sku': 'CT-001',
                    'notes': 'Exact match found'
                },
                {
                    'term': 'Barricade',
                    'quantity': 25,
                    'unit': 'EA',
                    'confidence': 0.88,
                    'matched_sku': 'CT-002',
                    'notes': 'Type III specification matched'
                }
            ],
            'alerts': [
                {
                    'type': 'Warning',
                    'message': 'Low stock on safety vests',
                    'severity': 'Medium',
                    'recommendation': 'Consider alternative suppliers'
                }
            ]
        },
        'pricing_summary': {
            'subtotal': 3312.50,
            'tax_rate': 0.085,
            'tax_amount': 281.56,
            'shipping': 150.00,
            'handling': 75.00,
            'total': 3819.06
        },
        'notes': 'All items meet CalTrans specifications. Delivery within 2 weeks guaranteed.'
    }


if __name__ == "__main__":
    # Example usage
    try:
        generator = ExcelBidGenerator("Zenlytic Solutions")
        sample_data = create_sample_bid_data()
        
        # Generate Excel file
        excel_bytes = generator.create_professional_bid(sample_data)
        
        # Save to file for testing
        with open('sample_bid.xlsx', 'wb') as f:
            f.write(excel_bytes)
        
        print("Sample bid document created successfully: sample_bid.xlsx")
        
    except Exception as e:
        print(f"Error creating sample bid: {e}") 